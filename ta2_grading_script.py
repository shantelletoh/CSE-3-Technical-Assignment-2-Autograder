import os
from bs4 import BeautifulSoup, Comment
import re
import requests
import base64
import csv
import pandas as pd
from openpyxl import load_workbook

# check for correct html file type
def check_type_html(file_name):
    global student_feedback
    global note_to_grader
    global total_score
    
    if not file_name.lower().endswith(".html"):
        student_feedback += "Error: The file is not an HTML file.\n"
        student_feedback += "Please email your correct assignment file to your TA so it can be graded."
        note_to_grader += "\nError: The file is not an HTML file."
        total_score -= 1000
        return

    with open(file_name, "r", encoding="utf-8") as fh:
        html_content = fh.read()
        # print(f"File: {file_name}\n")
        soup = BeautifulSoup(html_content, "html.parser")
        return soup

    
# 1. Title (this is different than a heading) (50 points)
def check_title():
    global student_feedback
    global total_score
    
    title_tag = soup.find("title")
    if title_tag is None:
        student_feedback += "-50: The file does not have a title.\n"
        total_score -= 50
        

# 2. Headings of two different sizes (ex. h1, h2, h3, etc.) (50 points each)
def check_headings():
    global student_feedback
    global total_score
    
    headings = soup.find_all(["h1", "h2", "h3", "h4", "h5", "h6"])
    heading_sizes = set()
    for heading in headings:
        heading_sizes.add(heading.name)
    # print(heading_sizes)

    if len(heading_sizes) == 1:
        student_feedback += "-50: The file contains only headings of one size (instead of two different sizes): " + ", ".join(heading_sizes) + ".\n"
        total_score -= 50
        
    elif len(heading_sizes) < 1:
        student_feedback += "-100: Does not have least two headings of different sizes.\n"
        total_score -= 100

# 3. Add a tooltip (or "hover box") to your first use of the h1 tag. (50 points)
def check_tooltip():
    global student_feedback
    global total_score
    
    h1_tag = soup.find("h1")
    if h1_tag:
        nested_tags_with_title = h1_tag.find_all(lambda tag: tag.has_attr("title"))
        tooltip = h1_tag.get("title")
        h1_class = h1_tag.get("class")
        if tooltip or len(nested_tags_with_title) > 0:
            return
        if h1_class and any("tooltip" in cls.lower() for cls in h1_class):
            return
        else:
            # find all <style> tags to extract CSS
            style_tags = soup.find_all("style")
            css_styles = []
            for style_tag in style_tags:
                css_styles.append(style_tag.get_text())

            # check if CSS contains styles for h1:hover or .tooltip class
            for css_style in css_styles:
                if "h1:hover" in css_style or ".tooltip" in css_style:
                    return
            student_feedback += "-50: The first <h1> tag in the file does not have a tooltip or hover box.\n"
            total_score -= 50
            return
    else:
        student_feedback += "-50: No <h1> tags.\n"
        total_score -= 50


# 4. Colored background (other than white) (100 points)
def check_colored_background():
    global student_feedback
    global total_score
    
    # colored background set in the CSS style tag
    style_tags = soup.find_all("style")
    for style_tag in style_tags:
        if "background-color" in style_tag.text and not "background-color: white" in style_tag.text:
            return

    # colored background also be set in the body tag
    body_tags = soup.find_all("body")
    if body_tags:
        for body_tag in body_tags:
            bgcolor = body_tag.get("bgcolor")
            if bgcolor:
                return
            elif body_tag.has_attr("style") and ("background-color" in body_tag["style"] or "background-image" in body_tag["style"]):
                return
        student_feedback += "-100: The file does not have a colored background.\n"
        total_score -= 100


# 5. "Mailto" link to your email address (50 points)
def check_mailto():
    global student_feedback
    global total_score
    
    a_tags = soup.find_all("a", href=True)
    for a_tag in a_tags:
        href = a_tag["href"]
        if href.lower().startswith("mailto:"):
            email_match = re.match(r"^mailto:([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})$", href, re.IGNORECASE)
            return
            # if email_match:
                # email_address = email_match.group(1)
                # print(f"Found email address in "{file_name}": {email_address}")
            # else:
                # print("No valid email address found.")
    student_feedback += "-50: The file does not have a \"mailto\" link to your email address.\n"
    total_score -= 50


# 6. Hyperlink to another website (must use a URL, not a file path) (50 points)
def check_hyperlink():
    global student_feedback
    global total_score
    
    anchor_tags = soup.find_all("a", href=True)

    for anchor_tag in anchor_tags:
        url = anchor_tag["href"]
        # Check if the URL is a valid HTTP or HTTPS link
        if re.match(r"^https?://", url):
            return
    student_feedback += "-50: No hyperlink to another website.\n"
    total_score -= 50


# 7. Bold text and italic text (50 points each)
def check_bold_and_italic():
    global student_feedback
    global total_score
    
    bold_tags = soup.find_all(["b", "strong"]) # Find all bold and strong tags
    italic_tags = soup.find_all(["i", "em"]) # Find all italic and emphasis tags
    
    bold = False
    italic = False
    
    # check inline styles
    elements_with_inline_bold_styles = soup.find_all(style=lambda value: value and ("font-weight:bold" in value or "font-weight: bold" in value))
    elements_with_inline_italic_styles = soup.find_all(style=lambda value: value and ("font-style:italic" in value or "font-style: italic" in value))
    
    # check style tag styles
    style_tags = soup.find_all("style")
    for tag in style_tags:
        if "font-weight:bold" in tag.text or "font-weight: bold" in tag.text:
            bold = True
            break
    for tag in style_tags:
        if "font-style:italic" in tag.text or "font-style: italic" in tag.text:
            italic = True
            break
            
    # print(not(bold_tags))
    # print(len(elements_with_inline_bold_styles) < 1)
    # print(bold)
    if (not(bold_tags)) and (len(elements_with_inline_bold_styles) < 1) and (bold is False):
        student_feedback += "-50: No bold text.\n"
        total_score -= 50

    if (not(italic_tags)) and (len(elements_with_inline_italic_styles) < 1) and (italic is False):
        student_feedback += "-50: No italic text.\n"
        total_score -= 50


# 8. Centered text or photo (50 points)
def check_centered_text_or_photo():
    global student_feedback
    global total_score
    
    style_tag = soup.find("style")
    if style_tag:
        style_text = style_tag.get_text()
        if "text-align: center" in style_text:
            return

    centered_elements_html = soup.find_all("center")
    centered_text_elements_css = soup.find_all(style=lambda value: value and "text-align: center" in value) # Find all elements with text-align: center; style
    centered_img_elements_css = soup.find_all("img", style=lambda value: value and "margin: auto" in value) # Find all img tags with margin: auto; style
    
    
    if not centered_elements_html and not centered_text_elements_css and not centered_img_elements_css:
        student_feedback += "-50: No centered content.\n"
        total_score -= 50


# 9. Horizontal line (aka "horizontal rule") (50 points)
def check_horizontal_line():
    global student_feedback
    global total_score
    
    horizontal_lines = soup.find_all("hr")
    if not horizontal_lines:
        student_feedback += "-50: No horizontal lines (horizontal rules).\n"
        total_score -= 50


# 10. Ordered list (numbered list) and unordered list (bulleted list) (50 points each)
def check_ordered_and_unordered_list():
    global student_feedback
    global total_score
    
    # li tags without parent count as unordered lists, so don't deduct points if there are such li tags
    check_unordered = True
    li_tags = soup.find_all("li")
    if len(li_tags) > 0:
        for li in li_tags:
            if li.find_parent(["ul", "ol"]) is None: # li tag has no ul or ol parent tag
                check_unordered = False
    
    ordered_lists = soup.find_all("ol")
    unordered_lists = soup.find_all("ul")
    if not ordered_lists:
        student_feedback += "-50: No ordered lists (numbered lists).\n"
        total_score -= 50

    if check_unordered and not unordered_lists:
        student_feedback += "-50: No unordered lists (bulleted lists).\n"
        total_score -= 50


# 11. A working picture, hosted online. You can link to an existing photo or upload your own image to photo-sharing sites like Google Photos or imgur.com. Make sure the photo is shared properly, and test your page on someone else's device to be certain. (100 points)
def check_picture():
    global student_feedback
    global note_to_grader
    global total_score
    
    picture_tags = soup.find_all("img")

    counter = 0
    # msg = False
    for picture_tag in picture_tags:
        counter += 1
        src = picture_tag.get("src")
        # print("src:", src)

        try:
            if src is None:
                student_feedback += "-100: No working picture hosted online.\n"
                total_score -= 100
                return
            
            # allow encoded images
            if src.startswith("data:image/"):
                encoded_data = re.search(r"base64,(.*)", src).group(1)
                base64.b64decode(encoded_data)
                return
                
            elif src.startswith("https://i.imgur.com/"): # special case so it won't give too many requests error
                return
            
            else:
                # try:
                response = requests.head(src)
                # print("src:", src)
                # print("response:", response)
                # print("response.status_code:", response.status_code)
                if response.status_code == 200:
                    # print("The file contains a working picture hosted online:")
                    # print("Image URL:", src)
                    return
                # elif response.status_code >= 300: # picture may or may not be visible
                #     msg = True
                elif response.status_code >= 300:
                    if counter == len(picture_tags): # last picture
                        note_to_grader += "\nCheck manually if there is an image. If image not visible, please manually deduct 100 points."
                        return
        except Exception as e:
            if counter == len(picture_tags): # last picture
                student_feedback += "-100: No working picture hosted online.\n"
                total_score -= 100
                return
    student_feedback += "-100: No working picture hosted online.\n"
    total_score -= 100


# 12. Set the width and height of your photo. Recommend 450x600 for portrait layout, or 600x450 for landscape. If your source photo isn't 4:3 aspect ratio, this may look odd. (100 points)
def check_picture_width_height():
    global student_feedback
    global total_score
    
    picture_tags = soup.find_all("img")

    for picture_tag in picture_tags:
        width = picture_tag.get("width")
        height = picture_tag.get("height")
        if width and height:
            return
        inline_style = picture_tag.get("style")
        if inline_style and (("width" in inline_style) or ("height" in inline_style)):
            return
    
    # find all <style> tags to extract CSS
    style_tags = soup.find_all("style")
    css_styles = []
    for style_tag in style_tags:
        css_styles.append(style_tag.get_text())

    # check if CSS contains width and height styles for img and image classes
    for css_style in css_styles:
        if "img" in css_style and ("width" in css_style or "height" in css_style):
            return
        if "image" in css_style and ("width" in css_style or "height" in css_style):
            return
    
    student_feedback += "-100: Either the image does not have width and height set, or there is no image.\n"
    total_score -= 100


# 13. Add a comment at the very bottom of your source code, listing the tools you used to create this: operating system, text editor, and the web browser you used to test it. (100 points)
def check_comment():
    global student_feedback
    global total_score
    
    comments = soup.find_all(string=lambda text: isinstance(text, Comment))
    for comment in comments:
        words = comment.strip().split()
        if len(words) >= 3:
            return

    comment_pattern = r"<!--\s*(?:\w+\W+){2,}"
    matches = soup.find_all(string=lambda text: re.search(comment_pattern, str(text)))
    if len(matches) > 0:
        return

    student_feedback += "-100: No comment for the tools used.\n"
    total_score -= 100


# Extra Credit: Add a video to your page, using YouTube, Vimeo, or other video hosting service. Find a video, look for the "Share" option, and select "Embed" or "Link" to find a working video link.
def check_extra_credit_video():
    global student_feedback
    global note_to_grader
    global total_score
    global extra_credit_done
    
    iframe_tags = soup.find_all("iframe")
    embed_tags = soup.find_all("embed")

    # check for video in iframe tag
    for iframe_tag in iframe_tags:
        src = iframe_tag.get("src")
        if src and ("youtube.com/embed" in src or "player.vimeo.com/video" in src):
            student_feedback += "+100 extra credit for embedding a video."
            # print("Video URL:", src)
            total_score += 100
            extra_credit_done = True
            return
        if src and ("youtube.com/watch" in src):
            student_feedback += "No extra credit: video is not correctly embedded."
            return
        else:
            note_to_grader += "\nAutograder cannot tell if video is correctly embedded. Please manually view the HTML file verify."
            return

    # check for video in embed tag
    for embed_tag in embed_tags:
        src = embed_tag.get("src")
        if src and ("youtube.com/embed" in src or "player.vimeo.com/video" in src):
            student_feedback += "+100 extra credit for embedding a video."
            # print("Video URL:", src)
            total_score += 100
            extra_credit_done = True
            return
        if src and ("youtube.com/watch" in src): 
            student_feedback += "No extra credit: video is not correctly embedded."
            return
        else:
            note_to_grader += "\nAutograder cannot tell if video is correctly embedded. Please manually view the HTML file verify."
            return

# append a note_to_grader to the appropriate cell in the csv for students who submitted multiple files
def multiple_files_note_to_grader(output_file):
    all_counts = []
    all_prev_values = []
    last_repeated_row_indices = []
    msgs = []
    
    with open(output_file, "r", newline="") as csvfile:
        reader = csv.reader(csvfile)
        
        count = 1 # keep track of number of files submitted
        prev_value = None
        for index, row in enumerate(reader):
            value = row[0] # column A
            if value == prev_value:
                count += 1
            elif count > 1:
                # print(f"{prev_value} submitted {count} files. Take the max score of the files submitted, and manually check submission if needed.")
                all_counts.append(count)
                all_prev_values.append(prev_value)
                last_repeated_row_indices.append(index)
                msgs.append(f"\n{prev_value} submitted {count} files. Take the max score of the files submitted, and manually check submission if needed.")
                count = 1
            prev_value = value
        
    
    with open(output_file, "r", newline="") as csvfile:
        reader = list(csv.reader(csvfile))
        
    for index, row in enumerate(last_repeated_row_indices):
        row_to_write_to = row - all_counts[index] # get the row of the student's 1st submitted file
        reader[row_to_write_to][3] += msgs[index] # add the note_to_grader to correct index in the list
    
    with open(output_file, "w", newline="") as csvfile: # rewrite csv file with the multiple submitted files msg
        writer = csv.writer(csvfile)
        writer.writerows(reader)


if __name__ == "__main__":
    
    output_file = "student_scores.csv"
    
    with open(output_file, "w", newline="") as csvfile:
        writer = csv.writer(csvfile)
        
        # write column headers to the csv
        column_headers = ["LastnameFirstname", "Total Score", "Student Feedback", "Note to TA/Grader"]
        writer.writerow(column_headers)

        folder_path = "./TA2_Submissions/"
        submissions_folder = sorted(os.listdir(folder_path)) # Get a list of all files in the folder in alphabetical order
        
        for file_name in submissions_folder:
            total_score = 1000 # max possible score
            extra_credit_done = False
            note_to_grader = "" # anything the grader should be aware of or needs to check
            student_feedback = "" # so students know why they got the score they got
            
            file_path = os.path.join(folder_path, file_name)
            
            print("file_path:", file_path)
            note_to_grader += "file_path: " + file_path
            parts = file_name.split("_")
            prefix = parts[0] # prefix (before 1st underscore) is lastnamefirstname of student
            
            # dock points for a late submission
            late = False
            if parts[1] == "LATE":
                student_feedback += "-100: late submission\n"
                note_to_grader += "\nLate submission. Please check the submission. If submitted within 1 hour after due date or penalty is excused, then remove the late penalty."
                total_score -= 100
            

            # check which requirements are met and which are not met
            soup = check_type_html(file_path)
            if soup:
                check_title() # 1.
                check_headings() # 2.
                check_tooltip() # 3.
                check_colored_background() # 4.
                check_mailto() # 5.
                check_hyperlink() # 6.
                check_bold_and_italic() # 7.
                check_centered_text_or_photo() # 8.
                check_horizontal_line() # 9.
                check_ordered_and_unordered_list() # 10.
                check_picture() # 11.
                check_picture_width_height() # 12.
                check_comment() # 13.
                check_extra_credit_video()
            
            
            # notes and feedback after done grading
            if total_score <= 700:
                note_to_grader += "\nTotal score is <= 700. Good idea to manually check the submission to ensure the grading script made no errors."
            elif total_score == 1100:
                student_feedback += "\nGood job! :)"
            elif total_score == 1000 and extra_credit_done is False:
                student_feedback += "Good job! :)"

            
            # write to csv file
            writer.writerow([prefix, total_score, student_feedback, note_to_grader])
    
    
    multiple_files_note_to_grader(output_file)
    
    
    
    df = pd.read_csv(output_file)
    
    # write dataframe to excel file
    df.to_excel("new_output_file.xlsx", index=False)
    
    
    # adjust excel column widths so the data fits
    wb = load_workbook("new_output_file.xlsx")
    
    for sheet in wb:
        for col in sheet.columns:
            column = col[0].column_letter # get column letter
            if column in ("A", "B"):
                max_length = 0
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                sheet.column_dimensions[column].width = max_length
                
    column_widths = {"C": 75, "D": 75} # manually adjust widths

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # set column widths based on the provided dictionary
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
    
    wb.save("new.xlsx")

    print("Grading script all done! Check the student_scores.csv file for the scores, student feedback, and notes to the grader.")